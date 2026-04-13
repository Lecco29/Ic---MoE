#!/usr/bin/env python3
# gera um arquivo xlsx com todos os resultados dos 3 experimentos
# pra abrir no google drive e converter pra planilhas google

import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ============================================================
# caminhos
# ============================================================
PASTA_RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_EXP1 = os.path.join(PASTA_RAIZ, 'experiments', 'exp01_camada_unica', 'results')
PASTA_EXP2 = os.path.join(PASTA_RAIZ, 'experiments', 'exp02_concatenacao', 'results')
PASTA_EXP3 = os.path.join(PASTA_RAIZ, 'experiments', 'exp03_moe', 'results')
ARQUIVO_SAIDA = os.path.join(PASTA_RAIZ, 'outputs', 'resultados_experimentos.xlsx')

# ============================================================
# cores (hex sem #)
# ============================================================
COR_TITULO      = 'FF1F3864'  # azul escuro
COR_CABECALHO   = 'FF2E75B6'  # azul medio
COR_SUBCABEC    = 'FFD6E4F0'  # azul claro
COR_MELHOR      = 'FFE2EFDA'  # verde claro (melhor resultado)
COR_PIOR        = 'FFFCE4D6'  # laranja claro (pior resultado)
COR_LINHA_PAR   = 'FFF5F5F5'  # cinza muito claro
COR_LINHA_IMPAR = 'FFFFFFFF'  # branco
COR_POSITIVO    = 'FF375623'  # verde escuro (melhora positiva)
COR_NEGATIVO    = 'FF9C0006'  # vermelho escuro (melhora negativa)

FONTE_TITULO    = Font(name='Calibri', bold=True, size=13, color='FFFFFFFF')
FONTE_CABEC     = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
FONTE_SUBCABEC  = Font(name='Calibri', bold=True, size=10, color='FF1F3864')
FONTE_NORMAL    = Font(name='Calibri', size=10)
FONTE_NEGRITO   = Font(name='Calibri', bold=True, size=10)

ALINHAMENTO_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALINHAMENTO_ESQ    = Alignment(horizontal='left',   vertical='center')

borda_fina = Side(style='thin', color='FFBFBFBF')
BORDA = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)


# ============================================================
# helpers
# ============================================================
def fill(cor):
    return PatternFill(fill_type='solid', fgColor=cor)


def escrever_celula(ws, linha, col, valor, fonte=None, preenchimento=None,
                    alinhamento=None, borda=True, formato=None):
    cel = ws.cell(row=linha, column=col, value=valor)
    if fonte:
        cel.font = fonte
    if preenchimento:
        cel.fill = preenchimento
    if alinhamento:
        cel.alignment = alinhamento
    else:
        cel.alignment = ALINHAMENTO_CENTRO
    if borda:
        cel.border = BORDA
    if formato:
        cel.number_format = formato
    return cel


def titulo_aba(ws, texto, num_colunas):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
    cel = ws.cell(row=1, column=1, value=texto)
    cel.font = FONTE_TITULO
    cel.fill = fill(COR_TITULO)
    cel.alignment = ALINHAMENTO_CENTRO
    cel.border = BORDA


def ajustar_larguras(ws, larguras):
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg


# ============================================================
# dados hardcoded dos experimentos (lidos dos CSVs)
# ============================================================

# Exp01 - melhor bloco/camada de cada backbone (para referencia)
# formato: backbone -> atributo -> [(camada, acc_mean, acc_std, f1)]
EXP01_IBOT_COLOR = [
    ('block0',  94.53, 1.13, 94.58),
    ('block1',  95.40, 1.27, 95.38),
    ('block2',  95.87, 1.05, 95.86),  # melhor
    ('block3',  95.20, 0.83, 95.19),
    ('block4',  92.13, 0.27, 92.19),
    ('block5',  88.73, 0.77, 88.78),
    ('block6',  84.93, 1.24, 84.89),
    ('block7',  84.07, 1.18, 83.95),
    ('block8',  76.00, 0.76, 75.82),
    ('block9',  69.93, 1.97, 69.81),
    ('block10', 66.13, 3.28, 65.94),
    ('block11', 62.80, 1.73, 62.66),
]
EXP01_IBOT_TEXTURE = [
    ('block0',  52.20, 2.69, 51.60),
    ('block1',  69.87, 2.28, 69.75),
    ('block2',  79.93, 2.61, 79.69),
    ('block3',  87.87, 1.44, 87.62),
    ('block4',  90.53, 1.33, 90.40),
    ('block5',  91.00, 2.02, 90.91),
    ('block6',  92.20, 1.85, 92.07),
    ('block7',  95.00, 0.99, 94.98),
    ('block8',  96.87, 0.54, 96.86),
    ('block9',  97.20, 0.54, 97.19),  # melhor
    ('block10', 96.27, 0.53, 96.25),
    ('block11', 96.00, 0.87, 95.97),
]
EXP01_RESNET_COLOR = [
    ('layer1', 95.47, 0.78, 95.48),
    ('layer2', 95.87, 0.50, 95.85),  # melhor
    ('layer3', 88.27, 1.12, 88.13),
    ('layer4', 59.20, 3.11, 59.19),
]
EXP01_RESNET_TEXTURE = [
    ('layer1', 83.40, 1.36, 83.39),
    ('layer2', 90.67, 1.65, 90.65),
    ('layer3', 93.27, 0.25, 93.23),  # melhor
    ('layer4', 86.33, 0.92, 86.23),
]
EXP01_VGG_COLOR = [
    ('layer1', 95.13, 0.81, 95.12),  # melhor
    ('layer2', 90.27, 1.27, 90.22),
    ('layer3', 85.47, 2.33, 85.35),
    ('layer4', 71.80, 2.55, 70.99),
    ('layer5', 55.07, 2.64, 55.12),
]
EXP01_VGG_TEXTURE = [
    ('layer1', 70.73, 2.25, 70.56),
    ('layer2', 86.60, 1.57, 86.61),
    ('layer3', 93.20, 1.48, 93.18),
    ('layer4', 94.73, 0.74, 94.76),  # melhor
    ('layer5', 89.40, 1.64, 89.42),
]
EXP01_VMAMBA_COLOR = [
    ('stage1', 94.41, 0.52, 94.35),  # melhor
    ('stage2', 93.41, 0.47, 93.33),
    ('stage3', 78.34, 3.75, 78.10),
    ('stage4', 54.14, 2.39, 53.94),
]
EXP01_VMAMBA_TEXTURE = [
    ('stage1', 93.86, 1.07, 93.33),
    ('stage2', 95.76, 0.84, 95.36),
    ('stage3', 95.76, 0.75, 95.37),  # melhor (empatado)
    ('stage4', 86.52, 2.57, 85.60),
]

# Exp02 - concatenacao
# formato: (backbone, atributo, camada_E, acc_E, std_E, f1_E, camada_D, acc_D, std_D, f1_D, acc_Z, std_Z, f1_Z, melhora)
EXP02 = [
    ('VGG16',    'Color',   'layer1', 95.13, 0.81, 95.12, 'layer4', 71.80, 2.55, 70.99, 86.13, 1.86, 85.96, -9.00),
    ('VGG16',    'Texture', 'layer1', 70.73, 2.25, 70.56, 'layer4', 94.73, 0.74, 94.76, 94.73, 0.90, 94.77, +0.00),
    ('ResNet50', 'Color',   'layer2', 95.87, 0.50, 95.85, 'layer3', 88.27, 1.12, 88.13, 94.13, 0.78, 94.08, -1.73),
    ('ResNet50', 'Texture', 'layer2', 90.67, 1.65, 90.65, 'layer3', 93.27, 0.25, 93.23, 93.53, 1.29, 93.51, +0.27),
    ('iBOT',     'Color',   'block2', 96.93, 0.49, 96.92, 'block9', 81.33, 1.62, 81.37, 85.13, 1.69, 85.12, -11.80),
    ('iBOT',     'Texture', 'block2', 78.33, 0.92, 78.11, 'block9', 87.80, 1.36, 87.73, 88.07, 1.04, 88.01, +0.27),
    ('VMamba',   'Color',   'stage1', 94.80, 0.62, 94.77, 'stage2', 93.93, 0.25, 93.90, 94.13, 0.16, 94.10, -0.67),
    ('VMamba',   'Texture', 'stage1', 88.40, 1.51, 88.21, 'stage2', 94.33, 1.40, 94.26, 93.40, 1.96, 93.29, -0.93),
]

# Exp03 - MoE (por fold + media final)
# formato: (backbone, atributo, [acc por fold], acc_mean, acc_std, f1_mean)
EXP03 = [
    ('iBOT',     'Color',   [96.67, 98.33, 97.33, 97.00, 97.33], 97.33, 0.56, 97.33),
    ('iBOT',     'Texture', [95.33, 95.00, 96.33, 96.67, 94.00], 95.47, 0.96, 95.45),
    ('ResNet50', 'Color',   [97.67, 97.00, 97.67, 96.00, 98.00], 97.27, 0.71, 97.25),
    ('ResNet50', 'Texture', [97.67, 97.00, 97.33, 98.33, 98.00], 97.67, 0.47, 97.65),
    ('VGG16',    'Color',   [94.00, 93.33, 93.00, 94.33, 92.33], 93.40, 0.71, 93.35),
    ('VGG16',    'Texture', [97.33, 94.67, 96.33, 97.00, 95.33], 96.13, 1.00, 96.12),
    ('VMamba',   'Color',   [96.33, 98.33, 96.00, 96.00, 98.00], 96.93, 1.02, 96.92),
    ('VMamba',   'Texture', [99.00, 97.00, 98.33, 97.33, 98.33], 98.00, 0.73, 98.00),
]


# ============================================================
# aba 1: exp01 - camada unica
# ============================================================
def criar_aba_exp01(wb):
    ws = wb.create_sheet(title='Exp01 - Camada Única')
    ws.freeze_panes = 'C4'

    titulo_aba(ws, 'Experimento 1 — Classificação por Camada Única  |  KNN (k=5), 5-fold, 70/30', 6)

    # cabecalho das colunas
    ws.row_dimensions[2].height = 8  # espaco
    cabecalhos = ['Backbone', 'Atributo', 'Camada', 'Acurácia Média (%)', '± Desvio Padrão', 'F1-Score (%)']
    for c, txt in enumerate(cabecalhos, start=1):
        escrever_celula(ws, 3, c, txt, fonte=FONTE_CABEC, preenchimento=fill(COR_CABECALHO))

    linha = 4
    todos_dados = [
        ('iBOT',     'Color',   EXP01_IBOT_COLOR,    max(EXP01_IBOT_COLOR,    key=lambda x: x[1])[0]),
        ('iBOT',     'Texture', EXP01_IBOT_TEXTURE,  max(EXP01_IBOT_TEXTURE,  key=lambda x: x[1])[0]),
        ('ResNet50', 'Color',   EXP01_RESNET_COLOR,  max(EXP01_RESNET_COLOR,  key=lambda x: x[1])[0]),
        ('ResNet50', 'Texture', EXP01_RESNET_TEXTURE,max(EXP01_RESNET_TEXTURE,key=lambda x: x[1])[0]),
        ('VGG16',    'Color',   EXP01_VGG_COLOR,     max(EXP01_VGG_COLOR,     key=lambda x: x[1])[0]),
        ('VGG16',    'Texture', EXP01_VGG_TEXTURE,   max(EXP01_VGG_TEXTURE,   key=lambda x: x[1])[0]),
        ('VMamba',   'Color',   EXP01_VMAMBA_COLOR,  max(EXP01_VMAMBA_COLOR,  key=lambda x: x[1])[0]),
        ('VMamba',   'Texture', EXP01_VMAMBA_TEXTURE,max(EXP01_VMAMBA_TEXTURE,key=lambda x: x[1])[0]),
    ]

    for backbone, atributo, camadas, melhor_camada in todos_dados:
        num = len(camadas)
        # merge backbone e atributo nas linhas do bloco
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha+num-1, end_column=1)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha+num-1, end_column=2)

        cel_back = ws.cell(row=linha, column=1, value=backbone)
        cel_back.font = FONTE_NEGRITO
        cel_back.fill = fill(COR_SUBCABEC)
        cel_back.alignment = ALINHAMENTO_CENTRO
        cel_back.border = BORDA

        cel_atr = ws.cell(row=linha, column=2, value=atributo)
        cel_atr.font = FONTE_NEGRITO
        cel_atr.fill = fill(COR_SUBCABEC)
        cel_atr.alignment = ALINHAMENTO_CENTRO
        cel_atr.border = BORDA

        for i, (camada, acc, std, f1) in enumerate(camadas):
            lin = linha + i
            cor = COR_MELHOR if camada == melhor_camada else (COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR)

            escrever_celula(ws, lin, 3, camada, fonte=FONTE_NORMAL, preenchimento=fill(cor))
            escrever_celula(ws, lin, 4, round(acc, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
            escrever_celula(ws, lin, 5, round(std, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
            escrever_celula(ws, lin, 6, round(f1, 2),  fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')

        linha += num + 1  # linha em branco entre backbones

    ws.row_dimensions[linha - 1].height = 6
    ajustar_larguras(ws, [12, 10, 10, 20, 18, 16])

    # legenda
    linha += 1
    escrever_celula(ws, linha, 1, '🟩 Melhor camada por backbone/atributo', fonte=FONTE_NORMAL,
                    preenchimento=fill(COR_MELHOR), borda=False, alinhamento=ALINHAMENTO_ESQ)


# ============================================================
# aba 2: exp02 - concatenacao
# ============================================================
def criar_aba_exp02(wb):
    ws = wb.create_sheet(title='Exp02 - Concatenação')
    ws.freeze_panes = 'C4'

    titulo_aba(ws, 'Experimento 2 — Fusão por Concatenação E+D  |  KNN (k=5), 5-fold, 70/30', 13)

    ws.row_dimensions[2].height = 8
    cabecalhos = [
        'Backbone', 'Atributo',
        'Camada E', 'Acc E (%)', '± Std', 'F1 E (%)',
        'Camada D', 'Acc D (%)', '± Std', 'F1 D (%)',
        'Acc Z (%)', '± Std', 'Melhora (%)'
    ]
    for c, txt in enumerate(cabecalhos, start=1):
        escrever_celula(ws, 3, c, txt, fonte=FONTE_CABEC, preenchimento=fill(COR_CABECALHO))

    # subcabecalhos visuais (linha separadora de grupos)
    grupos = [(1, 2, ''), (3, 6, 'Early (E)'), (7, 10, 'Deep (D)'), (11, 13, 'Fusão Z = E ⊕ D')]
    for inicio, fim, label in grupos:
        if label:
            ws.merge_cells(start_row=2, start_column=inicio, end_row=2, end_column=fim)
            cel = ws.cell(row=2, column=inicio, value=label)
            cel.font = FONTE_SUBCABEC
            cel.fill = fill(COR_SUBCABEC)
            cel.alignment = ALINHAMENTO_CENTRO
            cel.border = BORDA

    linha = 4
    for i, (back, atr, cam_e, acc_e, std_e, f1_e,
            cam_d, acc_d, std_d, f1_d, acc_z, std_z, f1_z, melhora) in enumerate(EXP02):
        cor = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR

        escrever_celula(ws, linha, 1, back,  fonte=FONTE_NEGRITO, preenchimento=fill(cor))
        escrever_celula(ws, linha, 2, atr,   fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        escrever_celula(ws, linha, 3, cam_e, fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        escrever_celula(ws, linha, 4, round(acc_e, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha, 5, round(std_e, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha, 6, round(f1_e,  2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha, 7, cam_d, fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        escrever_celula(ws, linha, 8, round(acc_d, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha, 9, round(std_d, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha,10, round(f1_d,  2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha,11, round(acc_z, 2), fonte=FONTE_NEGRITO,preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha,12, round(std_z, 2), fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')

        # celula de melhora: verde se positivo, vermelho se negativo
        cel_melhora = ws.cell(row=linha, column=13, value=round(melhora, 2))
        cel_melhora.number_format = '+0.00;-0.00;0.00'
        cel_melhora.border = BORDA
        cel_melhora.alignment = ALINHAMENTO_CENTRO
        cel_melhora.fill = fill(cor)
        if melhora > 0.05:
            cel_melhora.font = Font(name='Calibri', size=10, bold=True, color=COR_POSITIVO)
        elif melhora < -0.05:
            cel_melhora.font = Font(name='Calibri', size=10, bold=True, color=COR_NEGATIVO)
        else:
            cel_melhora.font = FONTE_NORMAL

        linha += 1

    ajustar_larguras(ws, [11, 9, 9, 11, 8, 10, 9, 11, 8, 10, 11, 8, 12])

    linha += 1
    nota = ws.cell(row=linha, column=1,
                   value='Melhora = Acc(Z) − max(Acc(E), Acc(D))  |  verde = concatenação ajudou  |  vermelho = concatenação prejudicou')
    nota.font = Font(name='Calibri', size=9, italic=True, color='FF595959')
    nota.alignment = ALINHAMENTO_ESQ


# ============================================================
# aba 3: exp03 - moe (por fold)
# ============================================================
def criar_aba_exp03_folds(wb):
    ws = wb.create_sheet(title='Exp03 - MoE (por fold)')
    ws.freeze_panes = 'C4'

    titulo_aba(ws, 'Experimento 3 — MoE: Mixture of Experts  |  KNN (k=5), 5-fold, 70/30', 10)

    ws.row_dimensions[2].height = 8
    cabecalhos = ['Backbone', 'Atributo',
                  'Fold 1 (%)', 'Fold 2 (%)', 'Fold 3 (%)', 'Fold 4 (%)', 'Fold 5 (%)',
                  'Média (%)', '± Std', 'F1 (%)']
    for c, txt in enumerate(cabecalhos, start=1):
        escrever_celula(ws, 3, c, txt, fonte=FONTE_CABEC, preenchimento=fill(COR_CABECALHO))

    linha = 4
    for i, (back, atr, folds, media, std, f1) in enumerate(EXP03):
        cor = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR

        escrever_celula(ws, linha, 1, back, fonte=FONTE_NEGRITO, preenchimento=fill(cor))
        escrever_celula(ws, linha, 2, atr,  fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        for j, acc_fold in enumerate(folds):
            escrever_celula(ws, linha, 3+j, round(acc_fold, 2),
                            fonte=FONTE_NORMAL, preenchimento=fill(cor), formato='0.00')
        escrever_celula(ws, linha, 8, round(media, 2), fonte=FONTE_NEGRITO, preenchimento=fill(COR_MELHOR), formato='0.00')
        escrever_celula(ws, linha, 9, round(std,   2), fonte=FONTE_NORMAL,  preenchimento=fill(cor),        formato='0.00')
        escrever_celula(ws, linha,10, round(f1,    2), fonte=FONTE_NORMAL,  preenchimento=fill(cor),        formato='0.00')

        linha += 1

    ajustar_larguras(ws, [11, 9, 10, 10, 10, 10, 10, 11, 8, 10])


# ============================================================
# aba 4: resumo comparativo
# ============================================================
def criar_aba_resumo(wb):
    ws = wb.create_sheet(title='Resumo Comparativo', index=0)
    ws.freeze_panes = 'C4'

    titulo_aba(ws, 'Resumo Comparativo — Exp01 vs Exp02 vs Exp03  |  Acurácia Média (%) por Backbone e Atributo', 8)

    ws.row_dimensions[2].height = 8
    cabecalhos = ['Backbone', 'Atributo',
                  'Exp01\nMelhor Camada', 'Exp01\nAcc (%)',
                  'Exp02\nFusão Z', 'Exp02\nAcc (%)',
                  'Exp03\nMoE', 'Exp03\nAcc (%)']
    for c, txt in enumerate(cabecalhos, start=1):
        escrever_celula(ws, 3, c, txt, fonte=FONTE_CABEC, preenchimento=fill(COR_CABECALHO))
    ws.row_dimensions[3].height = 30

    # melhor de cada exp01
    exp01_best = {
        ('iBOT',     'Color'):   ('block2',  95.87),
        ('iBOT',     'Texture'): ('block9',  97.20),
        ('ResNet50', 'Color'):   ('layer2',  95.87),
        ('ResNet50', 'Texture'): ('layer3',  93.27),
        ('VGG16',    'Color'):   ('layer1',  95.13),
        ('VGG16',    'Texture'): ('layer4',  94.73),
        ('VMamba',   'Color'):   ('stage1',  94.41),
        ('VMamba',   'Texture'): ('stage2',  95.76),
    }

    exp02_best = {
        ('VGG16',    'Color'):   86.13,
        ('VGG16',    'Texture'): 94.73,
        ('ResNet50', 'Color'):   94.13,
        ('ResNet50', 'Texture'): 93.53,
        ('iBOT',     'Color'):   85.13,
        ('iBOT',     'Texture'): 88.07,
        ('VMamba',   'Color'):   94.13,
        ('VMamba',   'Texture'): 93.40,
    }

    exp03_best = {
        ('iBOT',     'Color'):   97.33,
        ('iBOT',     'Texture'): 95.47,
        ('ResNet50', 'Color'):   97.27,
        ('ResNet50', 'Texture'): 97.67,
        ('VGG16',    'Color'):   93.40,
        ('VGG16',    'Texture'): 96.13,
        ('VMamba',   'Color'):   96.93,
        ('VMamba',   'Texture'): 98.00,
    }

    ordem = [
        ('iBOT',     'Color'),
        ('iBOT',     'Texture'),
        ('ResNet50', 'Color'),
        ('ResNet50', 'Texture'),
        ('VGG16',    'Color'),
        ('VGG16',    'Texture'),
        ('VMamba',   'Color'),
        ('VMamba',   'Texture'),
    ]

    linha = 4
    for i, (back, atr) in enumerate(ordem):
        cor = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR

        cam1, acc1 = exp01_best[(back, atr)]
        acc2 = exp02_best[(back, atr)]
        acc3 = exp03_best[(back, atr)]

        # descobre qual exp teve o melhor resultado
        melhor = max(acc1, acc2, acc3)

        def cor_cel(acc):
            return COR_MELHOR if abs(acc - melhor) < 0.01 else cor

        escrever_celula(ws, linha, 1, back, fonte=FONTE_NEGRITO, preenchimento=fill(cor))
        escrever_celula(ws, linha, 2, atr,  fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        escrever_celula(ws, linha, 3, cam1, fonte=FONTE_NORMAL,  preenchimento=fill(cor))
        escrever_celula(ws, linha, 4, round(acc1, 2),
                        fonte=FONTE_NEGRITO if cor_cel(acc1) == COR_MELHOR else FONTE_NORMAL,
                        preenchimento=fill(cor_cel(acc1)), formato='0.00')
        escrever_celula(ws, linha, 5, 'E ⊕ D', fonte=FONTE_NORMAL, preenchimento=fill(cor))
        escrever_celula(ws, linha, 6, round(acc2, 2),
                        fonte=FONTE_NEGRITO if cor_cel(acc2) == COR_MELHOR else FONTE_NORMAL,
                        preenchimento=fill(cor_cel(acc2)), formato='0.00')
        escrever_celula(ws, linha, 7, 'α·fE + β·fD', fonte=FONTE_NORMAL, preenchimento=fill(cor))
        escrever_celula(ws, linha, 8, round(acc3, 2),
                        fonte=FONTE_NEGRITO if cor_cel(acc3) == COR_MELHOR else FONTE_NORMAL,
                        preenchimento=fill(cor_cel(acc3)), formato='0.00')

        linha += 1

    ajustar_larguras(ws, [11, 9, 14, 12, 12, 12, 14, 12])

    linha += 2
    escrever_celula(ws, linha, 1, '🟩 Melhor resultado entre os 3 experimentos para cada backbone/atributo',
                    fonte=FONTE_NORMAL, preenchimento=fill(COR_MELHOR), borda=False, alinhamento=ALINHAMENTO_ESQ)

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)


# ============================================================
# main
# ============================================================
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba vazia padrao

    criar_aba_resumo(wb)
    criar_aba_exp01(wb)
    criar_aba_exp02(wb)
    criar_aba_exp03_folds(wb)

    os.makedirs(os.path.dirname(ARQUIVO_SAIDA), exist_ok=True)
    wb.save(ARQUIVO_SAIDA)
    print(f"Arquivo salvo em: {ARQUIVO_SAIDA}")
    print("Abra no Google Drive > botão direito > Abrir com > Planilhas Google")


if __name__ == '__main__':
    main()
